from openpyxl import load_workbook

class excel(object):	
	def __init__(self, filename, read_only, loadworkbook = load_workbook):
		self.filename=''
		self.read_only=True
		self.loadworkbook = loadworkbook

	def setWorkBook(self):
		self.wb = load_workbook(filename=self.filename, read_only=self.read_only)

	def setWorkSheet(self, sheet_name):
		self.ws = self.wb[sheet_name]

	def getRows(self):
		return self.ws.rows


